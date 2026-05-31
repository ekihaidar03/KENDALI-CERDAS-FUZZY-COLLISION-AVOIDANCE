import os
import ast
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


OUTPUT_DIR = "results/report_tables"
OUTPUT_FILE = "results/report_tables/laporan_simulasi_fuzzy_seano.xlsx"

SCENARIO_NAME = {
    "frontal_static_obstacle": "Obstacle frontal statis",
    "crossing_left_to_right": "Crossing kiri ke kanan",
    "crossing_right_to_left": "Crossing kanan ke kiri",
    "side_safe_obstacle": "Obstacle samping aman",
    "no_obstacle": "Tanpa obstacle",
}

COMMAND_COLUMNS = [
    "HOLD_COURSE",
    "SLOW_DOWN",
    "TURN_LEFT_SLOW",
    "TURN_RIGHT_SLOW",
    "STOP",
]

RISK_COLUMNS = [
    "LOW",
    "MEDIUM",
    "HIGH",
]


def buat_folder():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def parse_dict(teks):
    if pd.isna(teks):
        return {}

    if isinstance(teks, dict):
        return teks

    try:
        hasil = ast.literal_eval(str(teks))
        if isinstance(hasil, dict):
            return hasil
    except (ValueError, SyntaxError):
        pass

    return {}


def ambil_count(data_dict, key):
    try:
        return int(data_dict.get(key, 0))
    except (TypeError, ValueError):
        return 0


def nama_skenario(nama):
    return SCENARIO_NAME.get(nama, nama)


def status_validasi(row):
    scenario = row["scenario"]
    command_count = parse_dict(row.get("command_count", "{}"))

    risk_max = float(row.get("risk_max", 0.0))
    min_distance = row.get("min_distance_to_obstacle_m", None)

    if scenario == "no_obstacle":
        if ambil_count(command_count, "HOLD_COURSE") > 0:
            return "Valid - sistem mempertahankan HOLD_COURSE"
        return "Perlu evaluasi - command berubah saat tidak ada obstacle"

    if scenario == "side_safe_obstacle":
        if risk_max < 0.30 and ambil_count(command_count, "HOLD_COURSE") > 0:
            return "Valid - obstacle samping tidak memicu avoidance"
        return "Perlu evaluasi - obstacle samping memicu avoidance"

    if scenario == "frontal_static_obstacle":
        if ambil_count(command_count, "STOP") > 0:
            return "Valid - sistem melakukan STOP pada obstacle frontal"
        return "Perlu evaluasi - obstacle frontal belum menghasilkan STOP"

    if scenario in ["crossing_left_to_right", "crossing_right_to_left"]:
        if pd.notna(min_distance) and float(min_distance) > 0.30 and risk_max < 0.60:
            return "Valid simulatif - crossing terdeteksi sebagai risiko MEDIUM"
        return "Perlu evaluasi - jarak simulatif crossing terlalu kecil"

    return "Belum diklasifikasikan"


def catatan_analisis(row):
    scenario = row["scenario"]

    if scenario == "crossing_left_to_right":
        return (
            "Skenario crossing paling kritis. Fuzzy controller menurunkan kecepatan "
            "dan menjaga risk pada kelas MEDIUM. Jarak minimum simulatif perlu "
            "dibaca sebagai hasil model kinematik, bukan validasi fisik final."
        )

    if scenario == "crossing_right_to_left":
        return (
            "Obstacle crossing tetap berada pada kelas MEDIUM dan tidak mencapai "
            "kondisi STOP."
        )

    if scenario == "frontal_static_obstacle":
        return (
            "Obstacle frontal menghasilkan risk HIGH dan command STOP. Respons ini "
            "sesuai karakter konservatif collision avoidance."
        )

    if scenario == "side_safe_obstacle":
        return (
            "Obstacle berada di sisi luar corridor sehingga sistem tetap HOLD_COURSE."
        )

    if scenario == "no_obstacle":
        return (
            "Tidak ada obstacle. Sistem mempertahankan lintasan tanpa pergantian command."
        )

    return "-"


def load_data():
    path_fuzzy = "results/simulation_summary.csv"
    path_lintasan = "results/usv_trajectory_summary.csv"

    if not os.path.exists(path_fuzzy):
        raise FileNotFoundError(
            "File results/simulation_summary.csv belum ada. "
            "Jalankan dulu: python scenario_generator.py"
        )

    if not os.path.exists(path_lintasan):
        raise FileNotFoundError(
            "File results/usv_trajectory_summary.csv belum ada. "
            "Jalankan dulu: python usv_simulator.py"
        )

    fuzzy = pd.read_csv(path_fuzzy)
    lintasan = pd.read_csv(path_lintasan)

    return fuzzy, lintasan


def build_table_fuzzy(fuzzy):
    rows = []

    for _, row in fuzzy.iterrows():
        command_count = parse_dict(row.get("command_count", "{}"))
        risk_count = parse_dict(row.get("risk_class_count", "{}"))

        data = {
            "Skenario": nama_skenario(row["scenario"]),
            "Jumlah data": int(row["jumlah_data"]),
            "Risk minimum": float(row["risk_min"]),
            "Risk maksimum": float(row["risk_max"]),
            "Risk rata-rata": float(row["risk_mean"]),
            "Speed rata-rata": float(row["speed_mean"]),
        }

        for command in COMMAND_COLUMNS:
            data[command] = ambil_count(command_count, command)

        for risk in RISK_COLUMNS:
            data[risk] = ambil_count(risk_count, risk)

        rows.append(data)

    return pd.DataFrame(rows)


def build_table_lintasan(lintasan):
    rows = []

    for _, row in lintasan.iterrows():
        jarak = row["min_distance_to_obstacle_m"]

        rows.append(
            {
                "Skenario": nama_skenario(row["scenario"]),
                "Posisi akhir X (m)": float(row["final_x_m"]),
                "Posisi akhir Y (m)": float(row["final_y_m"]),
                "Deviasi lateral maksimum (m)": float(row["max_abs_cross_track_m"]),
                "Kecepatan rata-rata (m/s)": float(row["mean_speed_mps"]),
                "Kecepatan minimum (m/s)": float(row["min_speed_mps"]),
                "Jarak minimum simulatif (m)": None if pd.isna(jarak) else float(jarak),
                "Jumlah pergantian command": int(row["command_switch_count"]),
            }
        )

    return pd.DataFrame(rows)


def build_table_validasi(fuzzy, lintasan):
    gabung = fuzzy.merge(lintasan, on="scenario", how="left")
    rows = []

    for _, row in gabung.iterrows():
        command_count = parse_dict(row.get("command_count", "{}"))
        risk_count = parse_dict(row.get("risk_class_count", "{}"))

        jarak = row.get("min_distance_to_obstacle_m", None)

        data = {
            "Skenario": nama_skenario(row["scenario"]),
            "Risk maksimum": float(row["risk_max"]),
            "LOW": ambil_count(risk_count, "LOW"),
            "MEDIUM": ambil_count(risk_count, "MEDIUM"),
            "HIGH": ambil_count(risk_count, "HIGH"),
            "HOLD_COURSE": ambil_count(command_count, "HOLD_COURSE"),
            "SLOW_DOWN": ambil_count(command_count, "SLOW_DOWN"),
            "TURN_LEFT_SLOW": ambil_count(command_count, "TURN_LEFT_SLOW"),
            "TURN_RIGHT_SLOW": ambil_count(command_count, "TURN_RIGHT_SLOW"),
            "STOP": ambil_count(command_count, "STOP"),
            "Jarak minimum simulatif (m)": None if pd.isna(jarak) else float(jarak),
            "Status validasi simulasi": status_validasi(row),
            "Catatan analisis": catatan_analisis(row),
        }

        rows.append(data)

    return pd.DataFrame(rows)


def warna_header():
    return PatternFill("solid", fgColor="1F4E78")


def border_tipis():
    side = Side(style="thin", color="B7B7B7")
    return Border(left=side, right=side, top=side, bottom=side)


def tambah_judul(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=15, color="1F4E78")

    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="666666")
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")


def tulis_dataframe(ws, df, start_row=3, start_col=1):
    for j, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = warna_header()
        cell.border = border_tipis()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=start_col):
            value = row[col]

            if pd.isna(value):
                value = "-"

            cell = ws.cell(row=i, column=j, value=value)
            cell.border = border_tipis()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if isinstance(value, float):
                cell.number_format = "0.000"

    max_row = start_row + len(df)
    max_col = start_col + len(df.columns) - 1

    return start_row, start_col, max_row, max_col


def set_column_widths(ws, df):
    default_width = 14

    special_width = {
        "Skenario": 26,
        "Status validasi simulasi": 48,
        "Catatan analisis": 78,
        "Jarak minimum simulatif (m)": 22,
        "Deviasi lateral maksimum (m)": 24,
        "Jumlah pergantian command": 24,
        "Posisi akhir X (m)": 18,
        "Posisi akhir Y (m)": 18,
        "Kecepatan rata-rata (m/s)": 24,
        "Kecepatan minimum (m/s)": 24,
        "Risk maksimum": 16,
        "Risk minimum": 16,
        "Risk rata-rata": 16,
        "Speed rata-rata": 16,
    }

    for idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)

        if col_name in special_width:
            ws.column_dimensions[letter].width = special_width[col_name]
        elif col_name in COMMAND_COLUMNS:
            ws.column_dimensions[letter].width = 18
        elif col_name in RISK_COLUMNS:
            ws.column_dimensions[letter].width = 12
        else:
            ws.column_dimensions[letter].width = default_width


def set_row_heights(ws, min_row, max_row, text_heavy=False):
    ws.row_dimensions[min_row].height = 38

    for row in range(min_row + 1, max_row + 1):
        if text_heavy:
            ws.row_dimensions[row].height = 64
        else:
            ws.row_dimensions[row].height = 24


def format_table(ws, df, min_row, min_col, max_row, max_col, text_heavy=False):
    set_column_widths(ws, df)
    set_row_heights(ws, min_row, max_row, text_heavy=text_heavy)

    ws.freeze_panes = ws.cell(row=min_row + 1, column=min_col)
    ws.auto_filter.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border_tipis()


def beri_format_validasi(ws, min_row, max_row):
    header = {}
    for col in range(1, ws.max_column + 1):
        header[ws.cell(row=min_row, column=col).value] = col

    risk_col = header.get("Risk maksimum")
    jarak_col = header.get("Jarak minimum simulatif (m)")
    status_col = header.get("Status validasi simulasi")

    for row in range(min_row + 1, max_row + 1):
        if risk_col:
            risk_cell = ws.cell(row=row, column=risk_col)
            try:
                risk_value = float(risk_cell.value)

                if risk_value >= 0.60:
                    risk_cell.fill = PatternFill("solid", fgColor="F4CCCC")
                elif risk_value >= 0.30:
                    risk_cell.fill = PatternFill("solid", fgColor="FFF2CC")
                else:
                    risk_cell.fill = PatternFill("solid", fgColor="D9EAD3")
            except (TypeError, ValueError):
                pass

        if jarak_col:
            jarak_cell = ws.cell(row=row, column=jarak_col)
            try:
                jarak_value = float(jarak_cell.value)

                if jarak_value < 0.50:
                    jarak_cell.fill = PatternFill("solid", fgColor="F4CCCC")
                elif jarak_value < 1.00:
                    jarak_cell.fill = PatternFill("solid", fgColor="FFF2CC")
                else:
                    jarak_cell.fill = PatternFill("solid", fgColor="D9EAD3")
            except (TypeError, ValueError):
                pass

        if status_col:
            status_cell = ws.cell(row=row, column=status_col)
            status_text = str(status_cell.value)

            if "Perlu evaluasi" in status_text:
                status_cell.fill = PatternFill("solid", fgColor="F4CCCC")
            else:
                status_cell.fill = PatternFill("solid", fgColor="D9EAD3")


def buat_dashboard(wb, validasi):
    ws = wb.create_sheet("Dashboard")

    tambah_judul(
        ws,
        "Dashboard Simulasi Fuzzy Collision Avoidance USV SEANO",
        "Ringkasan hasil simulasi Mamdani Fuzzy Logic Controller berbasis parameter visual.",
    )

    total_skenario = len(validasi)
    jumlah_valid = validasi["Status validasi simulasi"].astype(str).str.contains("Valid").sum()
    jarak_min = validasi["Jarak minimum simulatif (m)"].dropna().min()
    risk_max = validasi["Risk maksimum"].max()

    indikator = [
        ["Jumlah skenario", total_skenario],
        ["Skenario valid simulatif", int(jumlah_valid)],
        ["Risk maksimum", float(risk_max)],
        ["Jarak minimum simulatif terkecil (m)", float(jarak_min)],
    ]

    ws["A4"] = "Indikator Utama"
    ws["A4"].font = Font(bold=True, size=12)

    for i, (nama, nilai) in enumerate(indikator, start=5):
        ws.cell(row=i, column=1, value=nama)
        ws.cell(row=i, column=2, value=nilai)

        ws.cell(row=i, column=1).border = border_tipis()
        ws.cell(row=i, column=2).border = border_tipis()

        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=2).number_format = "0.000"

    ws["A11"] = "Catatan Engineering"
    ws["A11"].font = Font(bold=True, size=12)

    ws["A12"] = (
        "Hasil jarak minimum merupakan nilai simulatif dari model kinematik sederhana. "
        "Nilai ini belum boleh disebut sebagai validasi keselamatan fisik final. "
        "Validasi fisik final tetap memerlukan pengujian langsung pada USV SEANO."
    )
    ws["A12"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A12:F14")

    ws["A16"] = "Rekomendasi Gambar untuk Laporan"
    ws["A16"].font = Font(bold=True, size=12)

    rekomendasi = [
        "G02_membership_visual_proximity.png",
        "G03_membership_approach_urgency.png",
        "G10_risk_frontal_static_obstacle.png",
        "G30_command_frontal_static_obstacle.png",
        "G40_trajectory_frontal_static_obstacle.png",
        "G40_trajectory_crossing_left_to_right.png",
        "G60_surface_risk_center_obstacle.png",
    ]

    for i, nama in enumerate(rekomendasi, start=17):
        ws.cell(row=i, column=1, value=nama)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    ws.row_dimensions[12].height = 54
    ws.freeze_panes = "A4"


def buat_sheet_tabel(wb, nama_sheet, title, df, validasi_style=False, text_heavy=False):
    ws = wb.create_sheet(nama_sheet)
    tambah_judul(ws, title)

    min_row, min_col, max_row, max_col = tulis_dataframe(ws, df, start_row=3, start_col=1)

    format_table(
        ws,
        df,
        min_row=min_row,
        min_col=min_col,
        max_row=max_row,
        max_col=max_col,
        text_heavy=text_heavy,
    )

    if validasi_style:
        beri_format_validasi(ws, min_row, max_row)

    return ws, min_row, max_row


def buat_sheet_catatan(wb):
    ws = wb.create_sheet("Catatan")

    tambah_judul(
        ws,
        "Catatan Penggunaan Hasil Simulasi",
        "Catatan ini disiapkan agar interpretasi hasil tidak berlebihan.",
    )

    catatan = [
        [
            "1",
            "Hasil ini merupakan validasi awal berbasis simulasi, bukan validasi keselamatan fisik final.",
        ],
        [
            "2",
            "Jarak minimum simulatif dihitung dari model kinematik sederhana pada program usv_simulator.py.",
        ],
        [
            "3",
            "Fitur proximity, area, bottom bounding box, dan vTTC berasal dari pendekatan visual kamera monocular sehingga tidak boleh diklaim sebagai jarak absolut.",
        ],
        [
            "4",
            "Validasi fisik final harus dilakukan pada USV SEANO melalui pengujian langsung, log runtime, video, dan observasi operator.",
        ],
        [
            "5",
            "Tabel ini digunakan untuk mendukung laporan Kendali Cerdas dan dokumentasi rancangan fuzzy controller.",
        ],
    ]

    ws["A4"] = "No"
    ws["B4"] = "Catatan"

    for cell in ["A4", "B4"]:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = warna_header()
        ws[cell].border = border_tipis()
        ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, row in enumerate(catatan, start=5):
        ws.cell(row=idx, column=1, value=row[0])
        ws.cell(row=idx, column=2, value=row[1])

        ws.cell(row=idx, column=1).border = border_tipis()
        ws.cell(row=idx, column=2).border = border_tipis()
        ws.cell(row=idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[idx].height = 42

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A5"


def buat_chart_command(ws, start_row, data_len):
    try:
        headers = [ws.cell(row=start_row, column=col).value for col in range(1, ws.max_column + 1)]

        command_start = headers.index("HOLD_COURSE") + 1
        command_end = headers.index("STOP") + 1
        scenario_col = headers.index("Skenario") + 1

        chart = BarChart()
        chart.title = "Distribusi Command per Skenario"
        chart.y_axis.title = "Jumlah"
        chart.x_axis.title = "Skenario"
        chart.height = 8.5
        chart.width = 23

        data = Reference(
            ws,
            min_col=command_start,
            max_col=command_end,
            min_row=start_row,
            max_row=start_row + data_len,
        )

        cats = Reference(
            ws,
            min_col=scenario_col,
            min_row=start_row + 1,
            max_row=start_row + data_len,
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend.position = "b"

        anchor_row = start_row + data_len + 5
        ws.add_chart(chart, f"A{anchor_row}")

    except Exception as exc:
        print(f"Chart tidak dibuat: {exc}")


def save_workbook(wb):
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.properties.creator = "Fuzzy Collision Avoidance Simulation"
    wb.properties.title = "Laporan Simulasi Fuzzy SEANO"
    wb.properties.subject = "Mamdani Fuzzy Logic Controller"
    wb.properties.created = datetime.now()

    try:
        wb.save(OUTPUT_FILE)
    except PermissionError:
        raise PermissionError(
            "File Excel output sedang terbuka. Tutup dulu "
            "results/report_tables/laporan_simulasi_fuzzy_seano.xlsx, lalu jalankan ulang."
        )


def main():
    buat_folder()

    fuzzy, lintasan = load_data()

    tabel_fuzzy = build_table_fuzzy(fuzzy)
    tabel_lintasan = build_table_lintasan(lintasan)
    tabel_validasi = build_table_validasi(fuzzy, lintasan)

    wb = Workbook()

    buat_dashboard(wb, tabel_validasi)

    ws_validasi, min_row_validasi, max_row_validasi = buat_sheet_tabel(
        wb,
        "Validasi Skenario",
        "Tabel Validasi Skenario Simulasi",
        tabel_validasi,
        validasi_style=True,
        text_heavy=True,
    )
    buat_chart_command(ws_validasi, start_row=min_row_validasi, data_len=len(tabel_validasi))

    buat_sheet_tabel(
        wb,
        "Ringkasan Fuzzy",
        "Ringkasan Output Fuzzy Controller",
        tabel_fuzzy,
        validasi_style=False,
        text_heavy=False,
    )

    buat_sheet_tabel(
        wb,
        "Ringkasan Lintasan",
        "Ringkasan Simulasi Lintasan USV",
        tabel_lintasan,
        validasi_style=False,
        text_heavy=False,
    )

    buat_sheet_catatan(wb)

    save_workbook(wb)

    print("\nWorkbook Excel laporan selesai dibuat.")
    print(f"File output: {OUTPUT_FILE}")
    print("\nSheet yang dibuat:")
    for sheet_name in wb.sheetnames:
        print(f"- {sheet_name}")


if __name__ == "__main__":
    main()